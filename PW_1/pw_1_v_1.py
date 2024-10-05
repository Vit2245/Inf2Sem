import openpyxl
from math import sin, cos,exp,log, fabs,pow
import numpy
# Открываем существующий файл Excel
#orkbook = openpyxl.load_workbook("data.xlsx")
try:
    # Попытка загрузить файл
    workbook = openpyxl.load_workbook("data.xlsx")

except FileNotFoundError:
    # Если файл не найден, создаем новый
    workbook = openpyxl.Workbook()
    # Создаем листы в файле формата *.xlsx
    workbook.create_sheet(index=0, title="Лист1")
    workbook.create_sheet(index=1, title="Лист2")
    workbook.create_sheet(index=2, title="Лист3")
    #сохранение файла
    workbook.save("data.xlsx")



sheet = workbook['Лист1']
sheet2 = workbook['Лист2']
sheet3 = workbook['Лист3']

# Задаем заголовки столбцов

def y(x):
    return sin(x)*exp(-2*x)


def g(x):
    if x<=0:
        return (1+pow(x,2))/pow(1+pow(x,4),(1/2))
    else:
        return 2*x+pow(sin(x),2)/(2+x)
def z(x):
    if x<-1:
        return (1+fabs(x))/pow(1+x+pow(x,2),3)
    elif x>0:
        return pow(1+x,3/5)
    else:
        return 2*log(1+pow(x,2))+(1+pow(cos(x),4))/(2+x)

def yy(x):
    return 2*sin(x)*cos(x)


def zz(x):
    if x<=0:
        return (1+pow(x,2))/pow(1+pow(x,4),(1/2))
    else:
        return 2*x+pow(sin(x),2)/(2+x)


def  clean(sheet,n,m):
    for i in range(2, len(n) + 2):
        # отчитска первых 5-ти столбиков
        for j in range(1, m+1):
            sheet.cell(row=i, column=j).value = None

def zzz(x,y):
    return pow(x,2)-2*pow(y,2)
# Определяем диапазон значений x
x_values = numpy.arange(-2.0, 2.1,0.1)
clean(sheet,x_values,3)
sheet['A1'] = 'x'
sheet['B1'] = 'y'
sheet['C1'] = 'g'
sheet['D1'] = 'z'
for i, x in enumerate(x_values, start=2):

    sheet.cell(row=i, column=1).value = x
    sheet.cell(row=i, column=2).value = y(x)
    sheet.cell(row=i, column=3).value = g(x)
    sheet.cell(row=i, column=4).value = z(x)
# Определяем диапазон значений x
x_values = numpy.arange(-2.0, 2.1, 0.1)
print(x_values)
clean(sheet2,x_values,3)
sheet2['A1'] = 'x'
sheet2['B1'] = 'y'
sheet2['C1'] = 'g'
sheet2['D1'] = 'z'
for i, x in enumerate(x_values, start=2):

    sheet2.cell(row=i, column=1).value = x
    sheet2.cell(row=i, column=2).value = yy(x)
    sheet2.cell(row=i, column=3).value = zz(x)

# Определяем диапазон значений x
x_values_3 = numpy.arange(-1.0, 1.1, 0.1)
y_values_3 = numpy.arange(-1.0, 1.1, 0.1)
print(x_values)
clean(sheet3,x_values,len(y_values_3))
sheet2['A1'] = 'x/y'
for i, x in enumerate(x_values_3, start=2):
    for j, y in enumerate(y_values_3, start=2):
        sheet3.cell(row=i, column=1).value = x
        sheet3.cell(row=1, column=j).value = y
        sheet3.cell(row=i, column=j).value = zzz(x,y)

# Сохраняем файл
workbook.save("data.xlsx")