import openpyxl
from math import sin, cos, exp, log, fabs, pow
import numpy

try:
    workbook = openpyxl.load_workbook("data.xlsx")
except FileNotFoundError:
    workbook = openpyxl.Workbook()
    workbook.create_sheet(index=0, title="Лист1")
    workbook.save("data.xlsx")

sheet = workbook['Лист1']

def y(x):
    return sin(x)*exp(-2*x)
def g(x):
    if x<=0:
        return (1+pow(x,2))/pow(1+pow(x,4),(1/2))
    else:
        return 2*x+pow(sin(x),2)/(2+x)
def z(x):
    if x<-1:
        return (1+fabs(x))/pow(1+x+pow(x,2),3)
    elif x>0:
        return pow(1+x,3/5)
    else:
        return 2*log(1+pow(x,2))+(1+pow(cos(x),4))/(2+x)

x_values = numpy.arange(-2.0, 2.1, 0.1)

# Заголовки столбцов (добавляем новые)
sheet['A1'] = 'x'
sheet['B1'] = 'y (for)'
sheet['C1'] = 'g (for)'
sheet['D1'] = 'z (for)'
sheet['E1'] = 'y (while)'
sheet['F1'] = 'g (while)'
sheet['G1'] = 'z (while)'

# Цикл for (остается без изменений)
for i, x in enumerate(x_values, start=2):
    sheet.cell(row=i, column=1).value = x
    sheet.cell(row=i, column=2).value = y(x)
    sheet.cell(row=i, column=3).value = g(x)
    sheet.cell(row=i, column=4).value = z(x)

# Цикл while (добавляем)
i = 2
j = 0
while j < len(x_values):
    x = x_values[j]
    sheet.cell(row=i, column=5).value = y(x) # Запись в столбцы E, F, G
    sheet.cell(row=i, column=6).value = g(x)
    sheet.cell(row=i, column=7).value = z(x)
    i += 1
    j += 1

workbook.save("data.xlsx")